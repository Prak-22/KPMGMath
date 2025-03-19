import streamlit as st
import os
import pandas as pd
import numpy as np
import pdfplumber
import pypdf
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import tempfile
import openai
import re
import base64

# Set page configuration
st.set_page_config(page_title="PDF to Excel Converter with Calculations", layout="wide")

# Check package versions
pd_version = pd.__version__
openpyxl_version = openpyxl.__version__

# Create output folder if it doesn't exist
if not os.path.exists("output_folder"):
    os.makedirs("output_folder")

# Sidebar for API key input and system info
with st.sidebar:
    st.title("Configuration")
    api_key = st.text_input("Enter your OpenAI API Key:", type="password")
    if api_key:
        # Store API key in session state instead of globally configuring openai
        st.session_state['openai_api_key'] = api_key

    # Check if API key is in session state
    if 'openai_api_key' in st.session_state:
        st.success("✅ API Key set")
    else:
        st.warning("⚠️ API Key not set")

    st.markdown("---")
    st.subheader("System Information")
    st.write(f"Pandas version: {pd_version}")
    st.write(f"Openpyxl version: {openpyxl_version}")

    # Check for potential version issues
    if pd_version >= "2.0.0" and openpyxl_version < "3.1.0":
        st.warning("Potential version incompatibility. Consider upgrading openpyxl:\n```\npip install openpyxl --upgrade\n```")

# Main application title
st.title("Financial Statement PDF to Excel Converter")
st.write("""
This application extracts tables from financial statement PDFs (like 10-K reports),
converts them to Excel, adds calculation rows to verify totals, and allows you to ask
questions about the financial data. It uses 100% Python libraries with no external dependencies.
""")

# Placeholder for the uploaded PDF content
pdf_content = None
extracted_tables = []
excel_path = None

# Function to handle duplicate column names (replacement for pd.io.parsers.ParserBase._maybe_dedup_names)
def handle_duplicate_columns(columns):
    """
    Creates unique column names by appending a suffix to duplicate column names.
    This replaces the functionality of pd.io.parsers.ParserBase._maybe_dedup_names
    which is no longer available in newer pandas versions.
    """
    new_columns = []
    seen = set()

    for i, col in enumerate(columns):
        original_col = col
        counter = 0

        while col in seen:
            counter += 1
            col = f"{original_col}.{counter}"

        seen.add(col)
        new_columns.append(col)

    return new_columns

# Function to extract tables from PDF
def extract_tables_from_pdf(pdf_file):
    tables = []

    # Save the uploaded file to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
        temp_file.write(pdf_file.getvalue())
        temp_path = temp_file.name

    # Extract tables using pdfplumber (Python-only solution)
    try:
        with st.spinner("Extracting tables using pdfplumber..."):
            with pdfplumber.open(temp_path) as pdf:
                total_pages = len(pdf.pages)
                progress_bar = st.progress(0)

                for page_num, page in enumerate(pdf.pages, 1):
                    progress_bar.progress(page_num / total_pages)
                    st.text(f"Processing page {page_num} of {total_pages}...")
                    extracted = page.extract_tables()
                    if extracted:
                        for table_num, table in enumerate(extracted, 1):
                            # Convert to pandas DataFrame
                            if table and len(table) > 1:  # Has at least headers and one row
                                # Get column names from the first row
                                columns = [str(col) if col is not None else f"Column_{i}"
                                        for i, col in enumerate(table[0])]

                                # Handle duplicate column names using our custom function
                                columns = handle_duplicate_columns(columns)

                                # Create DataFrame from remaining rows
                                df = pd.DataFrame(table[1:], columns=columns)

                                # Clean the dataframe
                                df = clean_financial_dataframe(df)

                                if not df.empty:
                                    tables.append(df)
                                    st.success(f"Successfully extracted Table {table_num} from Page {page_num}")
    except Exception as e:
        st.error(f"PDFPlumber extraction error: {str(e)}")

    # If no tables extracted with pdfplumber, try PyPDF
    if not tables:
        try:
            with st.spinner("Trying alternative extraction with PyPDF..."):
                with open(temp_path, 'rb') as f:
                    pdf_reader = pypdf.PdfReader(f)
                    total_pages = len(pdf_reader.pages)
                    progress_bar = st.progress(0)

                    for page_num in range(total_pages):
                        progress_bar.progress((page_num + 1) / total_pages)
                        st.text(f"Processing page {page_num + 1} of {total_pages} with PyPDF...")

                        # Extract text from page
                        page = pdf_reader.pages[page_num]
                        text = page.extract_text()

                        # Try to extract table-like structures from text
                        if text:
                            # Split by newlines to get rows
                            rows = text.split('\n')
                            processed_rows = []

                            for row in rows:
                                # Skip empty rows
                                if not row.strip():
                                    continue

                                # Try to split row by common delimiters
                                cells = re.split(r'\s{2,}', row)
                                if len(cells) > 1:  # Has multiple columns
                                    processed_rows.append(cells)

                            if len(processed_rows) > 1:  # At least headers and one data row
                                # First row as headers
                                headers = processed_rows[0]
                                # Make sure headers are unique using our custom function
                                headers = handle_duplicate_columns(headers)

                                # Create DataFrame
                                df = pd.DataFrame(processed_rows[1:], columns=headers)

                                # Clean the dataframe
                                df = clean_financial_dataframe(df)

                                if not df.empty:
                                    tables.append(df)
                                    st.success(f"Successfully extracted table from Page {page_num + 1} with PyPDF")
        except Exception as e:
            st.error(f"PyPDF extraction error: {str(e)}")

    # Clean up the temporary file
    try:
        os.unlink(temp_path)
    except:
        pass

    return tables

# Function to clean and prepare financial dataframes
def clean_financial_dataframe(df):
    # Remove completely empty rows and columns
    df = df.dropna(how='all').dropna(axis=1, how='all')

    if df.empty:
        return df

    # Handle duplicate column names
    df.columns = handle_duplicate_columns(df.columns)

    # If the first row looks like a header (contains 'assets', 'liabilities', etc.), use it as column names
    if len(df) > 0:
        first_row_text = ' '.join(df.iloc[0].astype(str)).lower()
        if any(keyword in first_row_text for keyword in ['assets', 'liabilities', 'equity', 'income', 'revenue', 'expenses']):
            # Create new column names from the first row, handling potential duplicates
            new_columns = [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(df.iloc[0])]
            new_columns = handle_duplicate_columns(new_columns)

            df.columns = new_columns
            df = df.iloc[1:]
            df = df.reset_index(drop=True)

    # Convert potential numeric columns to numeric types
    for col in df.columns:
        if col != df.columns[0]:  # Skip the first column which is usually labels
            # Try to convert numbers with commas, parentheses and $ signs
            try:
                # Remove $ and commas
                df[col] = df[col].astype(str).str.replace('$', '', regex=False).str.replace(',', '', regex=False)

                # Handle parentheses (negative numbers)
                df[col] = df[col].apply(lambda x: -float(x.strip('()')) if isinstance(x, str) and x.startswith('(') and x.endswith(')') else x)

                # Convert to numeric
                df[col] = pd.to_numeric(df[col], errors='coerce')
            except:
                pass

    # Rename columns if they are unnamed or numeric
    renamed_columns = []
    for i, col in enumerate(df.columns):
        if pd.isna(col) or (isinstance(col, (int, float)) and not isinstance(col, bool)):
            # If column is a date-looking number (like 20230729), try to format it
            if isinstance(col, (int, float)) and len(str(int(col))) == 8:
                date_str = str(int(col))
                try:
                    import datetime
                    date_obj = datetime.datetime.strptime(date_str, '%Y%m%d')
                    new_col = date_obj.strftime('%b %d, %Y')
                except:
                    new_col = f"Column_{i+1}"
            else:
                new_col = f"Column_{i+1}"
            renamed_columns.append(new_col)
        else:
            renamed_columns.append(col)

    # Handle potential duplicates in renamed columns
    renamed_columns = handle_duplicate_columns(renamed_columns)

    df.columns = renamed_columns

    return df

# Function to identify total rows
def identify_total_rows(df):
    total_rows = []

    # First, convert all column types to string to ensure we can search in them
    df_str = df.astype(str)

    # Keywords that indicate a total row
    total_keywords = [
        'total', 'subtotal', 'grand total', 'sum', 'totals',
        'total assets', 'total liabilities', 'total equity',
        'total current assets', 'total current liabilities'
    ]

    # Look through all rows
    for idx, row in df_str.iterrows():
        # Check first column specifically (often contains the total label)
        first_col = df_str.columns[0]
        first_cell = str(row[first_col]).lower()

        if any(keyword in first_cell for keyword in total_keywords):
            total_rows.append(idx)
            continue

        # If not found in first column, check all cells
        for col in df_str.columns:
            cell_value = str(row[col]).lower()
            if any(keyword in cell_value for keyword in total_keywords):
                total_rows.append(idx)
                break

    return total_rows

# Function to add calculation and difference rows
def add_calculation_rows(df, excel_writer, sheet_name):
    """
    Add calculation and difference rows to verify totals in the Excel file.

    Parameters:
    -----------
    df : pandas.DataFrame
        DataFrame containing the financial data
    excel_writer : pandas.ExcelWriter
        Excel writer object to write to
    sheet_name : str
        Name of the sheet to write to

    Returns:
    --------
    openpyxl.Workbook
        The processed workbook
    """
    try:
        # First, save the original dataframe to Excel
        df.to_excel(excel_writer, sheet_name=sheet_name, index=False)

        # Get the worksheet
        workbook = excel_writer.book
        worksheet = workbook[sheet_name]

        # Apply basic formatting to the entire worksheet
        apply_basic_formatting(worksheet)

        # Identify columns that are likely to contain numeric data
        numeric_cols = []
        for col_idx, col_name in enumerate(df.columns):
            if col_idx > 0:  # Skip the first column which is usually a label
                try:
                    # Try to convert the column to numeric
                    numeric_values = pd.to_numeric(df[col_name], errors='coerce')
                    # Check if column has any non-NA values after conversion
                    if numeric_values.notna().sum() > 0:
                        numeric_cols.append(col_idx + 1)  # +1 because Excel is 1-indexed
                except:
                    pass

        # Format numeric columns
        for col_idx in numeric_cols:
            format_numeric_column(worksheet, col_idx)

        # Identify total rows
        total_rows = identify_total_rows(df)

        # For each total row, add a calculation row and a difference row
        rows_to_add = 0
        for total_row_idx in total_rows:
            excel_row_idx = total_row_idx + 2 + rows_to_add  # +2 because Excel is 1-indexed and has a header row

            # Skip if this is beyond the worksheet boundaries
            if excel_row_idx > worksheet.max_row:
                continue

            # Get the value in the first column of the total row (usually contains "Total")
            total_label = worksheet.cell(row=excel_row_idx, column=1).value

            # Apply total row formatting
            apply_total_row_formatting(worksheet, excel_row_idx, len(df.columns))

            # Insert a calculation row
            worksheet.insert_rows(excel_row_idx + 1)
            calc_cell = worksheet.cell(row=excel_row_idx + 1, column=1)
            calc_cell.value = "Calculation"

            # Insert a difference row
            worksheet.insert_rows(excel_row_idx + 2)
            diff_cell = worksheet.cell(row=excel_row_idx + 2, column=1)
            diff_cell.value = "Difference"

            # Style for the new rows
            calc_fill = PatternFill(start_color="E6F1F5", end_color="E6F1F5", fill_type="solid")
            diff_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Apply styles to the calculation and difference rows
            apply_calculation_row_formatting(worksheet, excel_row_idx + 1, len(df.columns), calc_fill)
            apply_difference_row_formatting(worksheet, excel_row_idx + 2, len(df.columns), diff_fill)

            # Add calculations and formulas
            for col_idx in numeric_cols:
                # Determine the range for summation
                # Find previous total or start of data
                start_row = 2  # Default to first data row (after header)
                for prev_total_idx in reversed(total_rows):
                    if prev_total_idx < total_row_idx:
                        start_row = prev_total_idx + 2 + rows_to_add + 2  # +2 for Excel indexing, +2 for the added calc/diff rows
                        break

                # Range to sum (excluding previous calculation and difference rows)
                sum_range = []
                for r in range(start_row, excel_row_idx):
                    if r <= worksheet.max_row:  # Check if row exists
                        cell_value = worksheet.cell(row=r, column=1).value
                        if cell_value != "Calculation" and cell_value != "Difference":
                            sum_range.append(r)

                # If there are rows to sum
                if sum_range:
                    try:
                        # Create the SUM formula
                        cell_refs = [f"{get_column_letter(col_idx)}{r}" for r in sum_range]
                        sum_formula = f"=SUM({','.join(cell_refs)})"

                        # Add the calculation formula
                        calc_cell = worksheet.cell(row=excel_row_idx + 1, column=col_idx)
                        calc_cell.value = sum_formula
                        calc_cell.fill = calc_fill
                        calc_cell.number_format = '#,##0.00'

                        # Add the difference formula
                        diff_cell = worksheet.cell(row=excel_row_idx + 2, column=col_idx)
                        diff_formula = f"={get_column_letter(col_idx)}{excel_row_idx + 1}-{get_column_letter(col_idx)}{excel_row_idx}"
                        diff_cell.value = diff_formula
                        diff_cell.fill = diff_fill
                        diff_cell.number_format = '#,##0.00'
                    except Exception as e:
                        st.warning(f"Formula creation error: {str(e)}")

            # Update the counter for added rows
            rows_to_add += 2

        # Auto-adjust column widths
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Add a little extra space
            worksheet.column_dimensions[column].width = min(adjusted_width, 50)  # Cap width at 50

        return workbook
    except Exception as e:
        st.error(f"Error adding calculation rows: {str(e)}")
        # Return the workbook even if there was an error
        return excel_writer.book

# Helper function to apply basic formatting to the worksheet
def apply_basic_formatting(worksheet):
    # Format the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Add a thin border to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

# Helper function to format numeric columns
def format_numeric_column(worksheet, col_idx):
    for row_idx in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
        cell = worksheet.cell(row=row_idx, column=col_idx)
        try:
            value = cell.value
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        except:
            pass

# Helper function to apply total row formatting
def apply_total_row_formatting(worksheet, row_idx, num_cols):
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")

    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.font = total_font
        cell.fill = total_fill

# Helper function to apply calculation row formatting
def apply_calculation_row_formatting(worksheet, row_idx, num_cols, fill):
    calc_font = Font(italic=True)

    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.font = calc_font
        cell.fill = fill

# Helper function to apply difference row formatting
def apply_difference_row_formatting(worksheet, row_idx, num_cols, fill):
    diff_font = Font(italic=True)

    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.font = diff_font
        cell.fill = fill

        # Highlight cells with non-zero differences
        if col_idx > 1:  # Skip the first column (labels)
            # Add conditional formatting to highlight discrepancies
            cell.value = f"={get_column_letter(col_idx)}{row_idx-1}-{get_column_letter(col_idx)}{row_idx-2}"

            try:
                # Add formula to detect if difference is not zero
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Cell already has a formula, we'll use it
                    pass
                else:
                    # Create a formula to check for differences
                    cell.value = f"=IF(ABS({get_column_letter(col_idx)}{row_idx-1}-{get_column_letter(col_idx)}{row_idx-2})>0.01,{get_column_letter(col_idx)}{row_idx-1}-{get_column_letter(col_idx)}{row_idx-2},0)"
            except:
                pass

# Function to process the LLM query
def query_llm(question, tables_data):
    if not 'openai_api_key' in st.session_state or not st.session_state['openai_api_key']:
        return "Please enter your OpenAI API key in the sidebar to use the Q&A feature."

    # Prepare context from the tables
    context = "Here are the financial tables extracted from the document:\n\n"

    for i, df in enumerate(tables_data):
        context += f"Table {i+1}:\n{df.to_string()}\n\n"

    # Create the prompt for the LLM
    prompt = f"""
    You are a financial analyst assistant. Given the following financial data tables from a financial statement:

    {context}

    Please answer the following question:
    {question}

    Provide a clear, accurate, and concise answer based on the financial data provided.
    """

    # Define models to try in order of preference
    models_to_try = ["gpt-4", "gpt-3.5-turbo"]

    for model in models_to_try:
        try:
            # Create OpenAI client with just the API key
            client = openai.OpenAI(api_key=st.session_state['openai_api_key'])

            # Make the API call
            response = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": "You are a financial analysis assistant that helps interpret financial statements."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=500,
                temperature=0.3
            )

            # If we get here, the API call was successful
            model_used = f"(Using {model})"
            return f"{response.choices[0].message.content.strip()}\n\n{model_used if 'debug_mode' in st.session_state and st.session_state['debug_mode'] else ''}"

        except Exception as e:
            # If this is the last model to try, raise the exception
            if model == models_to_try[-1]:
                import traceback
                error_details = traceback.format_exc()
                return f"Error querying LLM: {str(e)}\n\nDetails: {error_details if 'debug_mode' in st.session_state and st.session_state['debug_mode'] else ''}"
            # Otherwise, try the next model
            continue

# Function to get a direct answer from the data (without LLM)
def get_direct_answer(question, tables_data):
    # Convert question to lowercase for easier matching
    question_lower = question.lower()

    # Parse date information from the question
    date_patterns = [
        r"july 29,? 2023",
        r"july 30,? 2022",
        r"2023",
        r"2022"
    ]

    target_date = None
    for pattern in date_patterns:
        if re.search(pattern, question_lower):
            target_date = pattern
            break

    # Look for total assets question
    if "total assets" in question_lower:
        return find_financial_value("total assets", target_date, tables_data)

    # Look for total current assets question
    elif "total current assets" in question_lower or "current assets" in question_lower:
        return find_financial_value("total current assets", target_date, tables_data)

    # Look for total liabilities question
    elif "total liabilities" in question_lower:
        return find_financial_value("total liabilities", target_date, tables_data)

    # Look for total equity question
    elif "total equity" in question_lower or "stockholders' equity" in question_lower:
        return find_financial_value("total equity", target_date, tables_data)

    # Look for matching assets and liabilities+equity question
    elif ("total assets" in question_lower and "total liabilities" in question_lower and
          ("equity" in question_lower or "match" in question_lower)):
        return check_balance_sheet_equality(tables_data)

    # For other types of questions, return None to indicate we should use the LLM
    return None

# Helper function to find specific financial values in tables
def find_financial_value(item_name, date_pattern, tables_data):
    for df in tables_data:
        # Convert DataFrame to string for easier searching
        df_str = df.astype(str)

        # Look for rows containing the item name
        for idx, row in df_str.iterrows():
            row_text = ' '.join(row.astype(str).values).lower()
            if item_name.lower() in row_text:
                # Find appropriate date column
                target_col = None
                if date_pattern:
                    for col in df.columns:
                        if date_pattern in str(col).lower():
                            target_col = col
                            break

                # If no specific date column found, take the first numeric column
                if not target_col:
                    for col in df.columns:
                        if col != df.columns[0]:  # Skip the first column (usually labels)
                            try:
                                if pd.to_numeric(df[col], errors='coerce').notna().any():
                                    target_col = col
                                    break
                            except:
                                pass

                if target_col:
                    try:
                        value = df.loc[idx, target_col]
                        date_str = date_pattern if date_pattern else "the most recent period"
                        return f"The value of {item_name} for {date_str} is {value}."
                    except:
                        pass

    return None

# Helper function to check if total assets equals total liabilities plus equity
def check_balance_sheet_equality(tables_data):
    assets_value_2023 = None
    liabilities_equity_2023 = None
    assets_value_2022 = None
    liabilities_equity_2022 = None

    # Search for values in all tables
    for df in tables_data:
        df_str = df.astype(str)

        # Look for total assets row
        for idx, row in df_str.iterrows():
            row_text = ' '.join(row.astype(str).values).lower()
            if "total assets" in row_text:
                # Find appropriate date columns
                col_2023 = None
                col_2022 = None

                for col in df.columns:
                    col_str = str(col).lower()
                    if "2023" in col_str:
                        col_2023 = col
                    elif "2022" in col_str:
                        col_2022 = col

                if col_2023 and assets_value_2023 is None:
                    try:
                        assets_value_2023 = pd.to_numeric(df.loc[idx, col_2023], errors='coerce')
                    except:
                        pass

                if col_2022 and assets_value_2022 is None:
                    try:
                        assets_value_2022 = pd.to_numeric(df.loc[idx, col_2022], errors='coerce')
                    except:
                        pass

            # Look for total liabilities and equity row
            elif "total liabilities and equity" in row_text or "total liabilities and stockholders' equity" in row_text:
                # Find appropriate date columns
                col_2023 = None
                col_2022 = None

                for col in df.columns:
                    col_str = str(col).lower()
                    if "2023" in col_str:
                        col_2023 = col
                    elif "2022" in col_str:
                        col_2022 = col

                if col_2023 and liabilities_equity_2023 is None:
                    try:
                        liabilities_equity_2023 = pd.to_numeric(df.loc[idx, col_2023], errors='coerce')
                    except:
                        pass

                if col_2022 and liabilities_equity_2022 is None:
                    try:
                        liabilities_equity_2022 = pd.to_numeric(df.loc[idx, col_2022], errors='coerce')
                    except:
                        pass

    # Generate response based on found values
    response = []

    if assets_value_2023 is not None and liabilities_equity_2023 is not None:
        if assets_value_2023 == liabilities_equity_2023:
            response.append(f"For 2023: Total Assets (${assets_value_2023}) matches Total Liabilities and Equity (${liabilities_equity_2023}).")
        else:
            response.append(f"For 2023: Total Assets (${assets_value_2023}) does NOT match Total Liabilities and Equity (${liabilities_equity_2023}).")

    if assets_value_2022 is not None and liabilities_equity_2022 is not None:
        if assets_value_2022 == liabilities_equity_2022:
            response.append(f"For 2022: Total Assets (${assets_value_2022}) matches Total Liabilities and Equity (${liabilities_equity_2022}).")
        else:
            response.append(f"For 2022: Total Assets (${assets_value_2022}) does NOT match Total Liabilities and Equity (${liabilities_equity_2022}).")

    if response:
        return "\n".join(response)
    else:
        return None

# Helper function to create a simple Excel file with formatting
def create_excel_file(tables, output_path):
    """
    Create an Excel file with formatted tables

    Parameters:
    -----------
    tables : list of pandas.DataFrame
        List of DataFrames to write to Excel
    output_path : str
        Path to save the Excel file

    Returns:
    --------
    str
        Path to the created Excel file
    """
    # Create an Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Process each table
        for i, df in enumerate(tables):
            sheet_name = f"Table_{i+1}"

            # Write dataframe to Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Format the worksheet
            workbook = writer.book
            worksheet = workbook[sheet_name]

            # Format headers
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            # Auto-adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column].width = min(adjusted_width, 50)

    return output_path

# File upload section
st.header("Step 1: Upload Financial Statement PDF")
uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")

if uploaded_file is not None:
    # Display PDF information
    st.success(f"Uploaded: {uploaded_file.name} ({round(uploaded_file.size/1024, 2)} KB)")
    pdf_content = uploaded_file

    # Extract button
    if st.button("Extract Tables from PDF"):
        with st.spinner("Extracting tables from PDF..."):
            extracted_tables = extract_tables_from_pdf(pdf_content)

            if extracted_tables:
                st.session_state['extracted_tables'] = extracted_tables
                st.success(f"Successfully extracted {len(extracted_tables)} tables from the PDF.")

                # Display tabs for each extracted table
                tabs = st.tabs([f"Table {i+1}" for i in range(len(extracted_tables))])

                for i, (tab, table) in enumerate(zip(tabs, extracted_tables)):
                    with tab:
                        st.dataframe(table)
            else:
                st.error("No tables were found in the PDF. Please try a different file or check if the PDF contains extractable tables.")

    # Generate Excel section
    if 'extracted_tables' in st.session_state and st.session_state['extracted_tables']:
        extracted_tables = st.session_state['extracted_tables']
        st.header("Step 2: Generate Excel with Calculations")

        # Create a simple debug toggle in the UI
        debug_mode = st.checkbox("Show Debug Information")
        simple_excel = st.checkbox("Use Simple Excel (No Calculations)", value=False)

        generate_excel = st.button("Generate Excel")
        if generate_excel:
            try:
                with st.spinner("Processing tables and generating Excel..."):
                    # Get a unique filename based on timestamp
                    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
                    base_filename = os.path.splitext(uploaded_file.name)[0]
                    output_filename = f"output_folder/{base_filename}_processed_{timestamp}.xlsx"

                    if debug_mode:
                        st.write(f"Creating Excel file: {output_filename}")
                        st.write(f"Number of tables to process: {len(extracted_tables)}")

                    if simple_excel:
                        # Use simple Excel generation (no calculations)
                        create_excel_file(extracted_tables, output_filename)
                    else:
                        # Use advanced Excel generation with calculations
                        # Create Excel file directly
                        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                            # Process each table
                            for i, table in enumerate(extracted_tables):
                                sheet_name = f"Table_{i+1}"
                                if debug_mode:
                                    st.write(f"Processing table {i+1} with shape {table.shape}")

                                # Add calculation rows
                                add_calculation_rows(table, writer, sheet_name)

                # Store in session state
                st.session_state['excel_path'] = output_filename

                st.success(f"Excel file generated successfully: {output_filename}")

                # Check if the file exists and is readable
                if os.path.exists(output_filename) and os.path.getsize(output_filename) > 0:
                    file_size_kb = os.path.getsize(output_filename)/1024
                    st.info(f"Excel file size: {file_size_kb:.2f} KB")

                    # Read the file for download
                    with open(output_filename, "rb") as file:
                        excel_bytes = file.read()

                    # Download button
                    st.download_button(
                        label=f"Download Excel File ({file_size_kb:.1f} KB)",
                        data=excel_bytes,
                        file_name=f"{base_filename}_processed.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key='download_excel'
                    )

                    # Show direct file path
                    if debug_mode:
                        st.code(f"File path: {os.path.abspath(output_filename)}")
                else:
                    st.error(f"Excel file was not created correctly. Check the output folder: {output_filename}")

            except Exception as e:
                st.error(f"Error generating Excel file: {str(e)}")
                import traceback
                if debug_mode:
                    st.error(traceback.format_exc())

                st.info("""
                **Troubleshooting Tips:**
                1. Make sure all dependencies are installed correctly: `pip install -r requirements.txt`
                2. Try updating openpyxl: `pip install openpyxl --upgrade`
                3. Check that you have write permissions to the output folder
                4. Try using the "Simple Excel" option if the normal option fails
                """)

        # Question answering section
        if 'excel_path' in st.session_state and os.path.exists(st.session_state['excel_path']):
            st.header("Step 3: Ask Questions About the Data")

            question = st.text_input("Enter your question about the financial data:")

            if question:
                if st.button("Get Answer"):
                    with st.spinner("Processing your question..."):
                        # Try to get a direct answer first
                        direct_answer = get_direct_answer(question, extracted_tables)

                        if direct_answer:
                            st.info(direct_answer)
                        else:
                            # If no direct answer, use LLM
                            if 'openai_api_key' in st.session_state and st.session_state['openai_api_key']:
                                answer = query_llm(question, extracted_tables)
                                st.info(answer)
                            else:
                                st.warning("Please enter your OpenAI API key in the sidebar to use the Q&A feature.")

# Instructions
with st.expander("How to Use This App"):
    st.write("""
    ### Instructions:

    1. **Upload a PDF**: Start by uploading a financial statement PDF (like a 10-K report).

    2. **Extract Tables**: Click the 'Extract Tables' button to process the PDF and extract tables.

    3. **Generate Excel**: After tables are extracted, click 'Generate Excel' to create an Excel file with:
       - The original table data
       - Added 'Calculation' rows that sum up line items
       - 'Difference' rows that show the difference between calculated totals and reported totals

    4. **Download the Excel**: Use the download button to save the Excel file to your computer.

    5. **Ask Questions**: Enter questions about the financial data to get answers based on the extracted tables.

    ### Dependencies:

    To run this app locally, install the required packages with:
    ```
    pip install streamlit openai pandas tabula-py pdfplumber openpyxl
    ```

    ### OpenAI API Key:

    To use the question answering feature, you need to enter your OpenAI API key in the sidebar.
    You can get an API key at https://platform.openai.com/account/api-keys
    """)

st.sidebar.markdown("---")
st.sidebar.info("""
    Created with:
    - Streamlit
    - OpenAI
    - Pandas
    - Tabula-py
    - PDFPlumber
    - Openpyxl
""")
